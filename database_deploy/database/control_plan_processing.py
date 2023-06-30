import pandas as pd
import pymssql
import backend.const
import openpyxl as opxl


## function push control plan from database admin/globalL
## PUSH FROM SHEET
def pushControlPlan(worksheet, size, type):
    cnxn = pymssql.connect(server= backend.const.jsonConst()["HOST"], port=backend.const.jsonConst()["PORT"]\
                        , database=backend.const.jsonConst()["DB_GET"]\
                        , user="new_sa"\
                        , password=backend.const.jsonConst()["PASSWORD"])
    if cnxn != None: print("PROCESS DATABASE: CONNECT SUCCESS", cnxn)
    cursor = cnxn.cursor()
    if type == "A":
        cursor.execute("SELECT * FROM " + backend.const.jsonConst["TABLE_CONTROL_PLAN_A"])
        for index in range(7, size):
            temp0 = str(worksheet.cell(row=index, column=1).value)
            temp1 = str(worksheet.cell(row=index, column=3).value) 
            temp2 = str(worksheet.cell(row=index, column=4).value)
            temp3 = str(worksheet.cell(row=index, column=5).value)
            temp4 = str(worksheet.cell(row=index, column=6).value)
            temp5 = str(worksheet.cell(row=index, column=15).value)
            temp6 = str(worksheet.cell(row=index, column=20).value)
            if temp0 == None: temp0 = '0'
            if temp1 == None: temp1 = '0'
            if temp2 == None: temp2 = '0'
            if temp3 == None: temp3 = '0'
            if temp4 == None: temp4 = '0'
            if temp5 == None: temp5 = '0'
            if temp6 == None: temp6 = '0'
            cursor.execute("""INSERT INTO """ + backend.const.jsonConst["TABLE_CONTROL_PLAN_A"] + """ (ITEM, CATALOG_NAME, PARA_NAME, POR_VALUE ,PRIORITY_VALUE, TAG_NAME, GU_ID) VALUES (%s, %s, %s, %s, %s, %s, %s)""", (temp0, temp1, temp2, temp3, temp4, temp5, temp6))
            cnxn.commit()
    elif type == "B":
        cursor.execute("SELECT * FROM " + backend.const.jsonConst["TABLE_CONTROL_PLAN_B"])
        for index in range(7, size):
            temp0 = str(worksheet.cell(row=index, column=1).value)
            temp1 = str(worksheet.cell(row=index, column=3).value) 
            temp2 = str(worksheet.cell(row=index, column=4).value)
            temp3 = str(worksheet.cell(row=index, column=5).value)
            temp4 = str(worksheet.cell(row=index, column=6).value)
            temp5 = str(worksheet.cell(row=index, column=15).value)
            temp6 = str(worksheet.cell(row=index, column=20).value)
            if temp0 == None: temp0 = '0'
            if temp1 == None: temp1 = '0'
            if temp2 == None: temp2 = '0'
            if temp3 == None: temp3 = '0'
            if temp4 == None: temp4 = '0'
            if temp5 == None: temp5 = '0'
            if temp6 == None: temp6 = '0'
            cursor.execute("""INSERT INTO """ + backend.const.jsonConst["TABLE_CONTROL_PLAN_B"] + """ (ITEM, CATALOG_NAME, PARA_NAME, POR_VALUE ,PRIORITY_VALUE, TAG_NAME, GU_ID) VALUES (%s, %s, %s, %s, %s, %s, %s)""", (temp0, temp1, temp2, temp3, temp4, temp5, temp6))
            cnxn.commit()
    elif type == "C":
        cursor.execute("SELECT * FROM " + backend.const.jsonConst["TABLE_CONTROL_PLAN_C"])
        for index in range(7, size):
            temp0 = str(worksheet.cell(row=index, column=1).value)
            temp1 = str(worksheet.cell(row=index, column=3).value) 
            temp2 = str(worksheet.cell(row=index, column=4).value)
            temp3 = str(worksheet.cell(row=index, column=5).value)
            temp4 = str(worksheet.cell(row=index, column=6).value)
            temp5 = str(worksheet.cell(row=index, column=15).value)
            temp6 = str(worksheet.cell(row=index, column=20).value)
            if temp0 == None: temp0 = '0'
            if temp1 == None: temp1 = '0'
            if temp2 == None: temp2 = '0'
            if temp3 == None: temp3 = '0'
            if temp4 == None: temp4 = '0'
            if temp5 == None: temp5 = '0'
            if temp6 == None: temp6 = '0'
            cursor.execute("""INSERT INTO """ + backend.const.jsonConst["TABLE_CONTROL_PLAN_C"] + """ (ITEM, CATALOG_NAME, PARA_NAME, POR_VALUE ,PRIORITY_VALUE, TAG_NAME, GU_ID) VALUES (%s, %s, %s, %s, %s, %s, %s)""", (temp0, temp1, temp2, temp3, temp4, temp5, temp6))
            cnxn.commit()
    cnxn.close()


## Get data from control plan to trackback
## GET TO DATAFRAME
def controlPlanGet(type):
    cnxn = pymssql.connect(server= backend.const.jsonConst()["HOST"], port=backend.const.jsonConst()["PORT"]\
                        , database=backend.const.jsonConst()["DB_GET"]\
                        , user="new_sa"\
                        , password=backend.const.jsonConst()["PASSWORD"])
    if cnxn != None: print("PROCESS DATABASE: CONNECT SUCCESS", cnxn)
    if type =="A":
        cur = cnxn.cursor()
        cur.execute("SELECT * FROM " + backend.const.TABLE_CONTROL_PLAN_A)
        data = cur.fetchall()
        df = pd.DataFrame(data)
        df['TOOL_VALUE']= None
    elif type =="B":
        cur = cnxn.cursor()
        cur.execute("SELECT * FROM " + backend.const.TABLE_CONTROL_PLAN_B)
        data = cur.fetchall()
        df = pd.DataFrame(data)
        df['TOOL_VALUE']= None
    elif type =="C":
        cur = cnxn.cursor()
        cur.execute("SELECT * FROM " + backend.const.TABLE_CONTROL_PLAN_C)
        data = cur.fetchall()
        df = pd.DataFrame(data)
        df['TOOL_VALUE']= None
    cnxn.close()
    return df

def readXLSX(name):
    xl = pd.ExcelFile(name+'.xlsx')
    data = pd.read_excel(xl, 0, header=None)
    return data

# read workbook
def readExcel(file_name):
    workbook = opxl.load_workbook(file_name)
    return workbook

# read worksheet
def readSheet(workbook):
    worksheet = workbook['5M+E Tool Parameter P3']
    return worksheet
##________________________________________________________________________________________________

if __name__ == "__main__":
    workbook = readExcel('output/laser_ce.xlsx')
    worksheet = readSheet(workbook)
    pushControlPlan(worksheet, 377, 'A')
    pushControlPlan(worksheet, 377, 'B')
    pushControlPlan(worksheet, 377, 'C')
    
    #python control_plan_processing.py